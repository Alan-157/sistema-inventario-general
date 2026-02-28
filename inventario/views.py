from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.urls import reverse
import json
from django.core.paginator import Paginator # Para asegurar que se descuente todo o nada
from django.db import models, transaction
from .models import Producto, HistorialMovimiento, Categoria, Proveedor, Pedido, DetallePedido
from django.contrib import messages
from usuarios.models import Usuario
from django.db.models import F, Count, Q # Importante para comparar campos
from django.core.exceptions import PermissionDenied
from django.http import HttpResponse,HttpResponseBadRequest, JsonResponse
from django.views.decorators.http import require_POST
from django.template.loader import get_template, render_to_string
from .forms import ProductoForm, MovimientoForm, CategoriaForm, ProveedorForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from xhtml2pdf import pisa
from django.utils import timezone
import datetime
from weasyprint import HTML
import tempfile

# --- SEGURIDAD Y DECORADORES ---

def es_admin_o_super(user):
    """Verifica si el usuario es ADMIN o SUPERUSUARIO según PPT U2-C4"""
    if user.rol in ['ADMIN', 'SUPERUSUARIO'] or user.is_superuser:
        return True
    raise PermissionDenied

@login_required
def dashboard(request):
    """Dashboard Dinámico: Cliente vs Trabajador vs Admin"""
    rol = request.user.rol # Definimos la variable para que no dé NameError
    
    if rol == 'CLIENTE':
        context = {
            'ultimos_productos': Producto.objects.filter(is_active=True, stock_actual__gt=0).order_by('-created_at')[:4],
            'total_disponibles': Producto.objects.filter(is_active=True, stock_actual__gt=0).count(),
            'categorias_activas': Categoria.objects.annotate(num_prods=Count('producto')).filter(num_prods__gt=0).count(),
            'mis_ultimos_pedidos': Pedido.objects.filter(cliente=request.user).order_by('-fecha_pedido')[:3]
        }
        return render(request, 'inventario/dashboard_cliente.html', context)
    
    if rol == 'TRABAJADOR':
        productos_criticos = Producto.objects.filter(is_active=True, stock_actual__lte=F('stock_minimo'))
        context = {
            'total_productos': Producto.objects.filter(is_active=True).count(),
            'cantidad_criticos': productos_criticos.count(),
            'productos_criticos': productos_criticos[:5],
            'ultimos_movimientos': HistorialMovimiento.objects.filter(is_active=True).select_related('producto', 'usuario').order_by('-created_at')[:8],
        }
        return render(request, 'inventario/dashboard_trabajador.html', context)

    # DASHBOARD ADMIN / SUPERUSUARIO
    productos_criticos = Producto.objects.filter(is_active=True, stock_actual__lte=F('stock_minimo'))
    datos_grafico = Categoria.objects.annotate(total=Count('producto')).filter(total__gt=0).values('nombre', 'total')
    
    labels_js = [item['nombre'] for item in datos_grafico]
    valores_js = [item['total'] for item in datos_grafico]

    context = {
        'total_productos': Producto.objects.filter(is_active=True).count(),
        'total_clientes': Usuario.objects.filter(rol='CLIENTE', is_active=True).count(),
        'pedidos_pendientes': Pedido.objects.filter(estado='PENDIENTE').count(),
        'cantidad_criticos': productos_criticos.count(),
        'productos_criticos': productos_criticos[:5],
        'ultimos_movimientos': HistorialMovimiento.objects.select_related('producto', 'usuario').order_by('-created_at')[:5],
        'labels_js': json.dumps(labels_js),
        'valores_js': json.dumps(valores_js),
    }
    return render(request, 'inventario/dashboard.html', context)

@login_required
def detalle_producto_json(request, pk):
    """Retorna los detalles de un producto para el modal AJAX"""
    producto = get_object_or_404(Producto, pk=pk)
    data = {
        'nombre': producto.nombre,
        'sku': producto.sku,
        'categoria': producto.categoria.nombre,
        'stock_actual': producto.stock_actual,
        'stock_minimo': producto.stock_minimo,
        'precio_venta': float(producto.precio_venta),
        'descripcion': producto.descripcion or "Sin descripción.",
        'imagen_url': producto.imagen.url if producto.imagen else None,
    }
    return JsonResponse(data)

@login_required
def lista_productos(request):
    # 1. Recolectar datos de los filtros (Igual que antes)
    query = request.GET.get('q', '')
    cat_id = request.GET.get('categoria', '')
    prov_id = request.GET.get('proveedor', '')
    precio_min = request.GET.get('min_precio', '')
    precio_max = request.GET.get('max_precio', '')

    # 2. Queryset base
    productos = Producto.objects.filter(is_active=True).select_related('categoria', 'proveedor').order_by('nombre')

    # 3. Aplicar Filtros
    if query:
        productos = productos.filter(Q(nombre__icontains=query) | Q(sku__icontains=query))
    if cat_id:
        productos = productos.filter(categoria_id=cat_id)
    if prov_id:
        productos = productos.filter(proveedor_id=prov_id)
    if precio_min:
        productos = productos.filter(precio_venta__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio_venta__lte=precio_max)

    # --- NUEVO: DETECTAR PETICIÓN AJAX ---
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        data = []
        for p in productos[:20]: # Limitamos a los primeros 20 para que vuele
            data.append({
                'id': p.id,
                'nombre': p.nombre,
                'sku': p.sku,
                'categoria': p.categoria.nombre,
                'stock': p.stock_actual,
                'stock_min': p.stock_minimo,
                'precio': p.precio_venta,
                'imagen_url': p.imagen.url if p.imagen else None,
                'url_detalle': reverse('detalle_producto', args=[p.id]),
                'url_editar': reverse('editar_producto', args=[p.id]),
                'url_desactivar': reverse('desactivar_producto', args=[p.id]),
            })
        return JsonResponse({'productos': data})

    # 4. Respuesta Normal (Paginación)
    paginator = Paginator(productos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    categorias = Categoria.objects.all()
    proveedores = Proveedor.objects.all()

    context = {
        'page_obj': page_obj,
        'categorias': categorias,
        'proveedores': proveedores,
        'query': query,
        'cat_seleccionada': cat_id,
        'prov_seleccionado': prov_id,
        'precio_min': precio_min,
        'precio_max': precio_max,
    }
    
    return render(request, 'inventario/lista_productos.html', context)

@login_required
@user_passes_test(es_admin_o_super)
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "¡Producto creado con éxito!")
            return redirect('lista_productos')
    else:
        form = ProductoForm()
    
    # Pasamos 'editando' en False para el título del template
    return render(request, 'inventario/crear_producto.html', {'form': form, 'editando': False})

@login_required
@user_passes_test(es_admin_o_super)
def editar_producto(request, pk):
    """Edita un producto existente (Solo Admin/Super)"""
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, f"Producto {producto.nombre} actualizado.")
            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'inventario/crear_producto.html', {'form': form, 'editando': True})

@login_required
def eliminar_producto(request, pk):
    if request.user.rol not in ['ADMIN', 'SUPERUSUARIO'] and not request.user.is_superuser:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'message': 'No tienes permisos'}, status=403)
        raise PermissionDenied
    
    producto = get_object_or_404(Producto, pk=pk)
    
    if request.method == 'POST':
        producto.is_active = False 
        producto.save()
        
        # Registrar en Kardex
        HistorialMovimiento.objects.create(
            producto=producto,
            usuario=request.user,
            tipo='SALIDA',
            cantidad=producto.stock_actual,
            motivo="BAJA LÓGICA: Producto desactivado"
        )
        
        # Si la petición es AJAX (SweetAlert)
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'message': f'El producto {producto.nombre} fue desactivado.'})
        
        messages.warning(request, f"El producto {producto.nombre} ha sido desactivado.")
        return redirect('lista_productos')
    
    return redirect('lista_productos')

@login_required
def productos_eliminados(request):
    if request.user.rol not in ['ADMIN', 'SUPERUSUARIO'] and not request.user.is_superuser:
        raise PermissionDenied

    # Filtramos solo los productos desactivados
    query = request.GET.get('q', '')
    productos = Producto.objects.filter(is_active=False)

    if query:
        productos = productos.filter(nombre__icontains=query) | productos.filter(sku__icontains=query)

    # Ordenar: por defecto los últimos desactivados primero
    productos = productos.order_by('-updated_at') # Asegúrate de tener auto_now=True en tu campo updated_at

    return render(request, 'inventario/productos_eliminados.html', {
        'productos': productos,
        'query': query
    })

@login_required
def reactivar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    producto.is_active = True
    producto.save()
    messages.success(request, f"El producto {producto.nombre} ha sido reactivado.")
    return redirect('productos_eliminados')

@login_required
def registrar_movimiento(request):
    # Capturamos el tipo desde la URL (?tipo=ENTRADA)
    tipo_predefinido = request.GET.get('tipo')
    
    if request.method == 'POST':
        form = MovimientoForm(request.POST)
        if form.is_valid():
            movimiento = form.save(commit=False)
            movimiento.usuario = request.user
            
            producto = movimiento.producto
            if movimiento.tipo == 'ENTRADA':
                producto.stock_actual += movimiento.cantidad
            elif movimiento.tipo == 'SALIDA':
                if producto.stock_actual >= movimiento.cantidad:
                    producto.stock_actual -= movimiento.cantidad
                else:
                    messages.error(request, f"Stock insuficiente de {producto.nombre}")
                    return render(request, 'inventario/registrar_movimiento.html', {'form': form})
            
            producto.save()
            movimiento.save()
            messages.success(request, "Movimiento registrado correctamente")
            return redirect('lista_movimientos')
    else:
        # Si hay tipo predefinido, lo pasamos como valor inicial al formulario
        initial_data = {}
        if tipo_predefinido in ['ENTRADA', 'SALIDA']:
            initial_data['tipo'] = tipo_predefinido
        form = MovimientoForm(initial=initial_data)

    return render(request, 'inventario/registrar_movimiento.html', {
        'form': form,
        'tipo_predefinido': tipo_predefinido
    })
    
    
@login_required
@user_passes_test(es_admin_o_super)
def editar_movimiento(request, pk):
    movimiento = get_object_or_404(HistorialMovimiento, pk=pk)
    producto = movimiento.producto
    cantidad_anterior = movimiento.cantidad
    tipo_anterior = movimiento.tipo

    if request.method == 'POST':
        form = MovimientoForm(request.POST, instance=movimiento)
        if form.is_valid():
            with transaction.atomic():
                # 1. Revertir el stock anterior
                if tipo_anterior == 'ENTRADA':
                    producto.stock_actual -= cantidad_anterior
                else:
                    producto.stock_actual += cantidad_anterior
                
                # 2. Aplicar el nuevo movimiento
                nuevo_movimiento = form.save(commit=False)
                if nuevo_movimiento.tipo == 'ENTRADA':
                    producto.stock_actual += nuevo_movimiento.cantidad
                else:
                    producto.stock_actual -= nuevo_movimiento.cantidad
                
                producto.save()
                nuevo_movimiento.save()
                messages.success(request, "Movimiento actualizado y stock reajustado.")
                return redirect('lista_movimientos')
    else:
        form = MovimientoForm(instance=movimiento)
    return render(request, 'inventario/registrar_movimiento.html', {'form': form, 'editando': True})

@login_required
def eliminar_movimiento(request, pk):
    # Solo ADMIN o SUPERUSUARIO pueden anular movimientos del Kardex
    if request.user.rol not in ['ADMIN', 'SUPERUSUARIO'] and not request.user.is_superuser:
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'message': 'No tienes permisos'}, status=403)
        raise PermissionDenied
    
    movimiento = get_object_or_404(HistorialMovimiento, pk=pk)
    
    if request.method == 'POST':
        producto = movimiento.producto
        
        # 1. REVERSIÓN DEL STOCK (Paso crítico)
        if movimiento.tipo == 'ENTRADA':
            # Si anulamos una entrada, restamos el stock que entró
            producto.stock_actual -= movimiento.cantidad
        elif movimiento.tipo == 'SALIDA':
            # Si anulamos una salida, devolvemos el stock que salió
            producto.stock_actual += movimiento.cantidad
        
        producto.save()

        # 2. DESACTIVACIÓN (Soft Delete)
        # Si tienes el campo is_active en HistorialMovimiento:
        movimiento.is_active = False 
        # Si no lo tienes, podemos cambiar el motivo para indicar la anulación
        movimiento.motivo = f"ANULADO por {request.user.username} - " + (movimiento.motivo or "")
        movimiento.save()
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({
                'ok': True, 
                'message': f'Movimiento anulado. El stock de {producto.nombre} ha sido corregido.'
            })
        
        messages.success(request, "Movimiento anulado correctamente.")
        return redirect('lista_movimientos')
    
    return redirect('lista_movimientos')

@login_required
def movimientos_anulados(request):
    if request.user.rol not in ['ADMIN', 'SUPERUSUARIO'] and not request.user.is_superuser:
        raise PermissionDenied

    # Filtramos solo los movimientos desactivados
    movimientos = HistorialMovimiento.objects.filter(is_active=False).select_related('producto', 'usuario').order_by('-updated_at')

    return render(request, 'inventario/movimientos_anulados.html', {
        'movimientos': movimientos
    })

@login_required
def restaurar_movimiento(request, pk):
    if request.user.rol not in ['ADMIN', 'SUPERUSUARIO'] and not request.user.is_superuser:
        raise PermissionDenied

    movimiento = get_object_or_404(HistorialMovimiento, pk=pk)
    producto = movimiento.producto

    if request.method == 'POST':
        # Al restaurar, debemos volver a afectar el stock
        if movimiento.tipo == 'ENTRADA':
            producto.stock_actual += movimiento.cantidad
        else:
            producto.stock_actual -= movimiento.cantidad
        
        producto.save()
        movimiento.is_active = True
        # Limpiamos el mensaje de "ANULADO" del motivo si lo pusiste
        movimiento.save()

        messages.success(request, f"Movimiento de {producto.nombre} restaurado. Stock actualizado.")
        return redirect('movimientos_anulados')
    
    return redirect('movimientos_anulados')

@login_required
def exportar_excel(request):
    """Exportación dinámica con auto-ajuste de columnas profesional"""
    
    # 1. Recuperar filtros de la URL (Hace que el excel sea dinámico)
    query = request.GET.get('q', '')
    cat_id = request.GET.get('categoria', '')
    prov_id = request.GET.get('proveedor', '')
    precio_min = request.GET.get('min_precio', '')
    precio_max = request.GET.get('max_precio', '')

    # 2. Queryset base
    productos = Producto.objects.filter(is_active=True).select_related('categoria', 'proveedor').order_by('nombre')

    # 3. Aplicar los mismos filtros que en la lista visual
    if query:
        productos = productos.filter(Q(nombre__icontains=query) | Q(sku__icontains=query))
    if cat_id:
        productos = productos.filter(categoria_id=cat_id)
    if prov_id:
        productos = productos.filter(proveedor_id=prov_id)
    if precio_min:
        productos = productos.filter(precio_venta__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio_venta__lte=precio_max)

    # 4. Crear el Libro y Estilos
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario"

    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_style = Alignment(horizontal="center", vertical="center")

    # 5. Escribir Encabezados y Datos
    headers = ['SKU', 'Producto', 'Categoría', 'Proveedor', 'Stock', 'Precio Venta', 'Valor Total']
    ws.append(headers)

    for p in productos:
        ws.append([
            p.sku, 
            p.nombre, 
            p.categoria.nombre, 
            p.proveedor.nombre if p.proveedor else "N/A",
            p.stock_actual,
            p.precio_venta,
            (p.stock_actual * p.precio_venta)
        ])

    # 6. Algoritmo de Auto-Ajuste de Columnas (Tu TOC solucionado)
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            try:
                if cell.value:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
            except:
                pass
        
        # Ajuste adaptable: largo + margen de respiro
        ws.column_dimensions[column_letter].width = max_length + 3

    # 7. Formatos de Celda (Moneda y Encabezado)
    for row in range(1, ws.max_row + 1):
        if row == 1:
            for cell in ws[row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_style
        else:
            # Columnas F y G como Moneda ($)
            ws.cell(row=row, column=6).number_format = '"$"#,##0'
            ws.cell(row=row, column=7).number_format = '"$"#,##0'
            # Columna E (Stock) centrada
            ws.cell(row=row, column=5).alignment = center_style

    # Activar filtros nativos de Excel
    ws.auto_filter.ref = ws.dimensions

    # 8. Respuesta de descarga
    response = HttpResponse(content_type='application/ms-excel')
    filename = f"Inventario_Filtrado_{datetime.datetime.now().strftime('%d-%m-%Y')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    
    return response

@login_required
def exportar_pdf(request):
    """Genera reporte de inventario en PDF"""
    productos = Producto.objects.filter(is_active=True).select_related('categoria')
    fecha_actual = datetime.datetime.now()
    
    context = {
        'productos': productos,
        'fecha': fecha_actual,
        'usuario': request.user
    }
    
    template = get_template('inventario/reporte_pdf.html')
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Inventario.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Error al generar PDF', status=500)
    return response

@login_required
@user_passes_test(es_admin_o_super)
def editar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    if request.method == 'POST':
        # request.FILES es clave para las imágenes (Clase 10)
        form = ProductoForm(request.POST, request.FILES, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, f"Producto {producto.nombre} actualizado correctamente.")
            return redirect('lista_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'inventario/crear_producto.html', {'form': form, 'editando': True})

@login_required
@user_passes_test(es_admin_o_super)
def desactivar_producto(request, pk):
    producto = get_object_or_404(Producto, pk=pk)
    # No borramos, desactivamos (Soft Delete)
    producto.is_active = False
    producto.save()
    messages.warning(request, f"Producto {producto.nombre} desactivado.")
    return redirect('lista_productos')

@login_required
def catalogo_cliente(request):
    # CAMBIA 'stock' por 'stock_actual'
    productos = Producto.objects.filter(stock_actual__gt=0, is_active=True)
    return render(request, 'inventario/catalogo_cliente.html', {'productos': productos})

@login_required
def detalle_producto(request, pk):
    """Muestra la ficha técnica de un producto (Normal o AJAX)"""
    producto = get_object_or_404(Producto, pk=pk, is_active=True)
    
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'nombre': producto.nombre,
            'sku': producto.sku,
            'categoria': producto.categoria.nombre,
            'proveedor': producto.proveedor.nombre if producto.proveedor else "N/A",
            'stock_actual': producto.stock_actual,
            'stock_minimo': producto.stock_minimo,
            'precio_costo': producto.precio_costo,
            'precio_venta': producto.precio_venta,
            'imagen_url': producto.imagen.url if producto.imagen else None,
            'created_at': producto.created_at.strftime("%d/%m/%Y"),
            'updated_at': producto.updated_at.strftime("%d/%m/%Y %H:%M"),
        })
    
    return render(request, 'inventario/detalle_producto.html', {'producto': producto})

@login_required
def lista_movimientos(request):
    """Muestra el historial completo con búsqueda AJAX y filtros de fecha/tipo"""
    if request.user.rol == 'CLIENTE':
        raise PermissionDenied

    # Captura de parámetros
    query = request.GET.get('q', '')
    tipo = request.GET.get('tipo', '')
    desde = request.GET.get('desde', '')
    hasta = request.GET.get('hasta', '')

    movimientos = HistorialMovimiento.objects.select_related('producto', 'usuario').order_by('-created_at')
    movimientos_list = HistorialMovimiento.objects.filter(is_active=True).select_related('producto', 'usuario').order_by('-created_at')
    
    # --- FILTROS ---
    if query:
        movimientos = movimientos.filter(
            Q(producto__nombre__icontains=query) |
            Q(producto__sku__icontains=query) |
            Q(motivo__icontains=query)
        )
    
    if tipo:
        movimientos = movimientos.filter(tipo=tipo)
        
    if desde and hasta:
        movimientos = movimientos.filter(created_at__date__range=[desde, hasta])

    # 1. RESPUESTA AJAX
    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        data = []
        for mov in movimientos[:30]: 
            data.append({
                'fecha': mov.created_at.strftime("%d/%m/%Y %H:%M"),
                'producto_nombre': mov.producto.nombre,
                'producto_sku': mov.producto.sku,
                'producto_id': mov.producto.pk,
                'tipo': mov.tipo,
                'cantidad': mov.cantidad,
                'usuario': mov.usuario.username,
                'motivo': mov.motivo or "Sin motivo especificado"
            })
        return JsonResponse({'movimientos': data})

    # 2. RESPUESTA NORMAL (Paginación)
    paginator = Paginator(movimientos, 15)
    page_obj = paginator.get_page(request.GET.get('page'))

    context = {
        'page_obj': page_obj, 
        'query': query,
        'tipo': tipo,
        'desde': desde,
        'hasta': hasta
    }
    return render(request, 'inventario/lista_movimientos.html', context)

from django.http import JsonResponse

@login_required
def detalle_movimiento(request, pk):
    # Buscamos el movimiento específico
    movimiento = get_object_or_404(HistorialMovimiento, pk=pk)
    
    # Preparamos los datos
    data = {
        'producto': movimiento.producto.nombre,
        'tipo': movimiento.get_tipo_display(),
        'cantidad': movimiento.cantidad,
        'usuario': movimiento.usuario.username,
        'fecha': movimiento.created_at.strftime("%d/%m/%Y %H:%M"),
        'motivo': movimiento.motivo or "Sin motivo registrado"
    }
    
    # Enviamos los datos de vuelta al navegador
    return JsonResponse(data)

@login_required
@user_passes_test(es_admin_o_super)
def lista_categorias(request):
    query = request.GET.get('q', '')
    
    # Obtenemos categorías y contamos cuántos productos tiene cada una
    categorias_list = Categoria.objects.annotate(
        num_productos=Count('producto')
    ).order_by('nombre')

    # Filtro de búsqueda
    if query:
        categorias_list = categorias_list.filter(
            Q(nombre__icontains=query) | Q(descripcion__icontains=query)
        )

    # Configuración del Paginador (5 por página)
    paginator = Paginator(categorias_list, 5)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventario/lista_categorias.html', {
        'page_obj': page_obj,
        'query': query
    })

@login_required
@user_passes_test(es_admin_o_super)
def crear_categoria(request):
    if request.method == 'POST':
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoría creada exitosamente.")
            return redirect('lista_categorias')
    else:
        form = CategoriaForm()
    return render(request, 'inventario/crear_categoria.html', {'form': form, 'editando': False})

@login_required
def detalle_categoria(request, pk):
    # Buscamos la categoría
    categoria = get_object_or_404(Categoria, pk=pk)
    
    # Contamos cuántos productos activos tiene esta categoría
    cantidad_productos = Producto.objects.filter(categoria=categoria, is_active=True).count()
    
    # Devolvemos la información en formato JSON
    return JsonResponse({
        'nombre': categoria.nombre,
        'descripcion': categoria.descripcion or "Sin descripción registrada.",
        'cantidad_productos': cantidad_productos,
    })

@login_required
@user_passes_test(es_admin_o_super)
def editar_categoria(request, pk):
    categoria = get_object_or_404(Categoria, pk=pk)
    if request.method == 'POST':
        form = CategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, f"Categoría '{categoria.nombre}' actualizada.")
            return redirect('lista_categorias')
    else:
        form = CategoriaForm(instance=categoria)
    return render(request, 'inventario/crear_categoria.html', {'form': form, 'editando': True})

@login_required
@require_POST
@user_passes_test(es_admin_o_super)
def eliminar_categoria(request, pk):
    if not request.headers.get("x-requested-with") == "XMLHttpRequest":
        return HttpResponseBadRequest("Solo AJAX")
    
    categoria = get_object_or_404(Categoria, pk=pk)
    
    # PROTECCIÓN: Si tiene productos, no se puede borrar
    if categoria.producto_set.count() > 0:
        return JsonResponse({
            "ok": False, 
            "message": "No puedes eliminar esta categoría porque tiene productos asociados."
        })
        
    nombre = categoria.nombre
    categoria.delete()
    return JsonResponse({"ok": True, "message": f"La categoría '{nombre}' ha sido eliminada."})

@login_required
@user_passes_test(es_admin_o_super)
def lista_proveedores(request):
    """Listado de proveedores con búsqueda y paginación real"""
    query = request.GET.get('q', '')
    
    # Obtenemos los proveedores y contamos sus productos
    proveedores_list = Proveedor.objects.annotate(
        num_productos=Count('producto')
    ).order_by('nombre')
    
    # Aplicamos filtro si existe búsqueda
    if query:
        proveedores_list = proveedores_list.filter(
            Q(nombre__icontains=query) | 
            Q(contacto__icontains=query)
        )
        
    # CONFIGURACIÓN DE PAGINACIÓN (Clase 9)
    # Mostraremos 5 por página para ver el efecto
    paginator = Paginator(proveedores_list, 5) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'inventario/lista_proveedores.html', {
        'page_obj': page_obj, 
        'query': query
    })
    
@login_required
def detalle_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    
    # Usamos getattr o verificamos el nombre real del campo en tu models.py
    # Si en tu modelo el campo es 'contacto', cambia 'nombre_contacto' por 'contacto'
    return JsonResponse({
        'nombre': proveedor.nombre,
        # Cambié 'nombre_contacto' por 'contacto' que es lo más probable que tengas
        'contacto': getattr(proveedor, 'contacto', "No registrado"), 
        'telefono': proveedor.telefono or "No registrado",
        'email': proveedor.email or "No registrado",
        'direccion': proveedor.direccion or "No registrada",
        'productos_suministrados': proveedor.producto_set.filter(is_active=True).count(),
    })

@login_required
@user_passes_test(es_admin_o_super)
def crear_proveedor(request):
    if request.method == 'POST':
        form = ProveedorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Proveedor registrado exitosamente.")
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm()
    return render(request, 'inventario/form_proveedor.html', {'form': form, 'editando': False})

@login_required
@user_passes_test(es_admin_o_super)
def editar_proveedor(request, pk):
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if request.method == 'POST':
        form = ProveedorForm(request.POST, instance=proveedor)
        if form.is_valid():
            form.save()
            messages.success(request, f"Proveedor {proveedor.nombre} actualizado.")
            return redirect('lista_proveedores')
    else:
        form = ProveedorForm(instance=proveedor)
    return render(request, 'inventario/form_proveedor.html', {'form': form, 'editando': True})

@login_required
@require_POST
@user_passes_test(es_admin_o_super)
def eliminar_proveedor(request, pk):
    if not request.headers.get("x-requested-with") == "XMLHttpRequest":
        return HttpResponseBadRequest()
    
    proveedor = get_object_or_404(Proveedor, pk=pk)
    if proveedor.producto_set.count() > 0:
        return JsonResponse({"ok": False, "message": "No se puede eliminar: tiene productos asociados."})
        
    proveedor.delete()
    return JsonResponse({"ok": True, "message": "Proveedor eliminado correctamente."})

@login_required
def crear_pedido(request):
    if request.method == 'POST':
        # Aquí simplificamos: el cliente envía una lista de IDs y cantidades
        producto_ids = request.POST.getlist('productos[]')
        cantidades = request.POST.getlist('cantidades[]')
        
        if producto_ids:
            nuevo_pedido = Pedido.objects.create(cliente=request.user)
            total_pedido = 0
            
            for p_id, cant in zip(producto_ids, cantidades):
                producto = Producto.objects.get(id=p_id)
                cantidad = int(cant)
                precio = producto.precio_venta
                
                DetallePedido.objects.create(
                    pedido=nuevo_pedido,
                    producto=producto,
                    cantidad=cantidad,
                    precio_unitario=precio
                )
                total_pedido += (precio * cantidad)
            
            nuevo_pedido.total = total_pedido
            nuevo_pedido.save()
            
            messages.success(request, "¡Pedido enviado con éxito! Espera la aprobación.")
            return redirect('mis_pedidos')
            
    return redirect('catalogo_cliente')

@login_required
def mis_pedidos(request):
    # 'detalles__producto' trae los nombres de los productos de un solo golpe
    pedidos = Pedido.objects.filter(cliente=request.user).prefetch_related('detalles__producto').order_by('-fecha_pedido')
    return render(request, 'inventario/mis_pedidos.html', {'pedidos': pedidos})

# Vista para el Administrador
@login_required
def gestion_pedidos(request):
    if request.user.rol not in ['ADMIN', 'SUPERUSUARIO'] and not request.user.is_superuser:
        raise PermissionDenied
    
    # Captura de filtros
    estado_filtro = request.GET.get('estado', '')
    cliente_filtro = request.GET.get('cliente', '')
    fecha_desde = request.GET.get('desde', '')
    fecha_hasta = request.GET.get('hasta', '')

    # Consulta base optimizada con select_related
    pedidos = Pedido.objects.all().select_related('cliente').order_by('-fecha_pedido')

    # --- Aplicación de Filtros ---
    if estado_filtro:
        pedidos = pedidos.filter(estado=estado_filtro)
    if cliente_filtro:
        pedidos = pedidos.filter(cliente__username__icontains=cliente_filtro)
    if fecha_desde and fecha_hasta:
        pedidos = pedidos.filter(fecha_pedido__date__range=[fecha_desde, fecha_hasta])

    # Paginación
    paginator = Paginator(pedidos, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'inventario/gestion_pedidos.html', {
        'pedidos': page_obj,
        'estado_actual': estado_filtro,
        'cliente_query': cliente_filtro,
        'desde': fecha_desde,
        'hasta': fecha_hasta,
    })
    
@login_required
def detalle_pedido(request, pk):
    pedido = get_object_or_404(Pedido, pk=pk)
    # Por ahora solo renderizamos, puedes crear el template 'detalle_pedido.html' después
    return render(request, 'inventario/detalle_pedido.html', {'pedido': pedido})

# Función para cambiar el estado (Aprobar/Rechazar)
@login_required
def cambiar_estado_pedido(request, pedido_id, nuevo_estado):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    # Seguridad básica
    if not request.user.is_superuser and request.user.rol not in ['ADMIN', 'SUPERUSUARIO']:
        messages.error(request, "No tienes permisos para esta acción.")
        return redirect('dashboard')

    # Solo procesamos si el pedido está PENDIENTE
    if pedido.estado == 'PENDIENTE':
        if nuevo_estado == 'APROBADO':
            try:
                with transaction.atomic(): # Si falla un producto, se cancela todo el proceso
                    for detalle in pedido.detalles.all():
                        producto = detalle.producto
                        if producto.stock_actual >= detalle.cantidad:
                            # Restamos del stock actual
                            producto.stock_actual -= detalle.cantidad
                            producto.save()
                        else:
                            # Error si no alcanza el stock
                            raise ValueError(f"Stock insuficiente para: {producto.nombre}")
                    
                    pedido.estado = 'APROBADO'
                    pedido.save()
                    messages.success(request, f"Pedido #{pedido.id} aprobado. Stock actualizado.")
            
            except ValueError as e:
                messages.error(request, str(e))
        
        elif nuevo_estado == 'RECHAZADO':
            pedido.estado = 'RECHAZADO'
            pedido.save()
            messages.warning(request, f"Pedido #{pedido.id} rechazado.")

    return redirect('gestion_pedidos')

@login_required
def aprobar_pedido(request, pedido_id):
    # Seguridad: Solo admin o superusuario
    if not request.user.is_superuser and request.user.rol not in ['ADMIN', 'SUPERUSUARIO']:
        messages.error(request, "No tienes permisos para aprobar pedidos.")
        return redirect('dashboard')

    pedido = get_object_or_404(Pedido, id=pedido_id)

    if pedido.estado == 'PENDIENTE':
        try:
            # transaction.atomic asegura que si falla el descuento de un producto, 
            # no se descuente nada de los otros (evita datos corruptos)
            with transaction.atomic():
                for detalle in pedido.detalles.all():
                    producto = detalle.producto
                    
                    if producto.stock_actual >= detalle.cantidad:
                        # 1. Descontar stock
                        producto.stock_actual -= detalle.cantidad
                        producto.save()

                        # 2. Crear movimiento en el Kardex automáticamente
                        HistorialMovimiento.objects.create(
                            producto=producto,
                            usuario=request.user, # El admin que aprueba
                            tipo='SALIDA',
                            cantidad=detalle.cantidad,
                            motivo=f"Venta/Pedido aprobado #{pedido.id}"
                        )
                    else:
                        raise ValueError(f"Stock insuficiente para {producto.nombre}")

                # 3. Marcar pedido como aprobado
                pedido.estado = 'APROBADO'
                pedido.save()
                messages.success(request, f"Pedido #{pedido.id} aprobado con éxito.")
                
        except ValueError as e:
            messages.error(request, str(e))
    
    return redirect('gestion_pedidos')

@login_required
def rechazar_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    if pedido.estado == 'PENDIENTE':
        pedido.estado = 'RECHAZADO'
        pedido.save()
        messages.warning(request, f"Pedido #{pedido.id} ha sido rechazado.")
    return redirect('gestion_pedidos')

@login_required
def generar_pdf_pedido(request, pedido_id):
    pedido = get_object_or_404(Pedido, id=pedido_id)
    
    # Seguridad: Solo el dueño del pedido o un admin puede descargarlo
    if pedido.cliente != request.user and not request.user.is_superuser:
        return HttpResponse("No autorizado", status=401)

    # Renderizamos un template HTML especial para el PDF
    html_string = render_to_string('inventario/pdf_pedido.html', {'pedido': pedido})
    
    # Creamos el objeto PDF
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    result = html.write_pdf()

    # Preparamos la respuesta del navegador
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Comprobante_Pedido_{pedido.id}.pdf"'
    response.write(result)
    
    return response

@login_required
def reporte_inventario_pdf(request):
    # Seguridad: Solo admin
    if not request.user.is_superuser and request.user.rol not in ['ADMIN', 'SUPERUSUARIO']:
        return HttpResponse("No autorizado", status=401)

    productos = Producto.objects.filter(is_active=True).order_by('categoria', 'nombre')
    
    # Calculamos el valor total del inventario (Precio costo * Stock)
    valor_total = sum(p.precio_costo * p.stock_actual for p in productos)

    html_string = render_to_string('inventario/pdf_inventario.html', {
        'productos': productos,
        'valor_total': valor_total,
        'fecha': timezone.now()
    })
    
    html = HTML(string=html_string, base_url=request.build_absolute_uri())
    result = html.write_pdf()

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Inventario_General.pdf"'
    response.write(result)
    
    return response

@login_required
def reporte_ventas_pdf(request):
    # Seguridad: Solo Admins
    if request.user.rol not in ['ADMIN', 'SUPERUSUARIO']:
        raise PermissionDenied

    # Solo consideramos pedidos APROBADOS para el reporte de ventas
    pedidos = Pedido.objects.filter(estado='APROBADO').order_by('-fecha_pedido')
    
    # Calculamos el total general de ventas
    total_ventas = sum(pedido.total for pedido in pedidos)
    
    # Datos adicionales para el reporte
    cantidad_pedidos = pedidos.count()
    promedio_venta = total_ventas / cantidad_pedidos if cantidad_pedidos > 0 else 0

    html_string = render_to_string('inventario/pdf_ventas.html', {
        'pedidos': pedidos,
        'total_ventas': total_ventas,
        'cantidad_pedidos': cantidad_pedidos,
        'promedio_venta': promedio_venta,
        'fecha': timezone.now()
    })
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Ventas_General.pdf"'
    
    # Usamos WeasyPrint para generar el PDF
    HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf(response)
    
    return response

@login_required
def eliminar_producto_logico(request, pk):
    if request.user.rol not in ['ADMIN', 'SUPERUSUARIO']:
        raise PermissionDenied
    
    producto = get_object_or_404(Producto, pk=pk)
    producto.is_active = False # Cambiamos el estado en lugar de borrar
    producto.save()
    
    # Registramos en el Kardex que se dio de baja
    HistorialMovimiento.objects.create(
        producto=producto,
        usuario=request.user,
        tipo='SALIDA',
        cantidad=producto.stock_actual,
        motivo="BAJA LÓGICA: Producto desactivado del sistema"
    )
    
    messages.warning(request, f"El producto {producto.nombre} ha sido desactivado.")
    return redirect('gestion_inventario')

@login_required
def gestion_inventario(request):
    """Gestión de Inventario con Filtros y Paginación"""
    query = request.GET.get('q', '')
    cat_id = request.GET.get('categoria', '')
    
    # 1. Queryset base (Solo activos)
    productos_list = Producto.objects.filter(is_active=True).select_related('categoria', 'proveedor').order_by('nombre')
    
    # 2. Aplicar filtros
    if query:
        productos_list = productos_list.filter(Q(nombre__icontains=query) | Q(sku__icontains=query))
    if cat_id:
        productos_list = productos_list.filter(categoria_id=cat_id)

    # 3. Paginación (10 por página)
    paginator = Paginator(productos_list, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Datos para select de filtros
    categorias = Categoria.objects.filter(is_active=True)

    return render(request, 'inventario/gestion_inventario.html', {
        'page_obj': page_obj, 
        'query': query,
        'categorias': categorias,
        'cat_seleccionada': cat_id
    })
    
@login_required
def detalle_categoria_json(request, pk):
    cat = get_object_or_404(Categoria, pk=pk)
    # Contamos cuántos productos tiene esta categoría
    prods_count = Producto.objects.filter(categoria=cat, is_active=True).count()
    return JsonResponse({
        'nombre': cat.nombre,
        'descripcion': cat.descripcion or "Sin descripción.",
        'total_productos': prods_count
    })

@login_required
def detalle_proveedor_json(request, pk):
    prov = get_object_or_404(Proveedor, pk=pk)
    return JsonResponse({
        'nombre': prov.nombre,
        'contacto': prov.contacto or "No registrado",
        'telefono': prov.telefono or "No registrado",
        'email': prov.email or "No registrado",
        'direccion': prov.direccion or "No registrada"
    })